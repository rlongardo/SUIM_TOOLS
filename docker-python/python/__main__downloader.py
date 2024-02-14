import os, sys, logging, subprocess, datetime, locale, re
import openpyxl
from config import get_connection, DB_CONNECTIONS, BRANCH_OFFICE, PACS_STUDIES, PACS_STUDIES_DXA

locale.setlocale(locale.LC_ALL, ("es_MX", "UTF-8"))
AET=os.environ.get('AET')
STATE=os.environ.get('STATE')
IP1= DB_CONNECTIONS[STATE].split()                                                                  #Se obtienen los valores de la conexión de DB y se divide en subcad
HOST= IP1[3][6:-1]
PORT=os.environ.get('PORT')

def patients_count(template_file):
    book = openpyxl.load_workbook(template_file, data_only=True)
    sheet = book.active
    count = 0
    for cell in sheet['A']:
        if isinstance(cell.value, str) and "AN" in cell.value:
            count += 1
    book.close()
    return count

def directory_create(sheet,i,directory_name):
    obj_accession = sheet.cell(row = i, column = 1)
    accessionNumber = obj_accession.value
    obj_patient_name = sheet.cell(row = i, column = 2)
    patient_name = obj_patient_name.value
    obj_patient_id = sheet.cell(row = i, column = 3)
    patient_id = obj_patient_id.value
    obj_images = sheet.cell(row = i, column = 5)
    images = obj_images.value
    obj_date = sheet.cell(row = i, column = 6)
    year = obj_date.value.strftime('%Y')
    month = obj_date.value.strftime('%m')
    day = obj_date.value.strftime("%d")
    obj_unidad = sheet.cell(row = i, column = 7)
    unidad = obj_unidad.value
    modality = sheet.cell(row = i, column = 4).value
    #patient_directory = os.path.join(directory_name, unidad, modality, year, month, day, patient_id + "-" + patient_name)
    patient_directory = os.path.join(directory_name, unidad, modality, year, month, day, patient_name)
    if not os.path.exists(directory_name):
        os.makedirs(directory_name)
    return patient_directory, accessionNumber, modality, patient_name, images

def dicom_rename(UID, accessionNumber,patient_directory):
    count = 0
    for path, dirs, files in os.walk(patient_directory):
        for file in files:
            count += 1
            #new_name = str(count) + "UN." + UID + "_" + accessionNumber + "_" + ".dcm"
            new_name = str(count) + "UN." + UID + ".dcm"
            os.rename(os.path.join(path,file), os.path.join(path,new_name))

def first_validation(images, patient_directory, accessionNumber, modality):                                  #Mejorar la comprobación
    count = 0
    for path, dirs, files in os.walk(patient_directory):
        count = len(files)
    if images == count:
        logging.info("Se observan DICOM existentes del ANC " + str(accessionNumber) + " correctos " + str(images) + "/" + str(count))
    elif count > images:
        logging.warning("El estudio " + str(accessionNumber) + " || Modalidad: " + modality + " debe contener " + str(images) + " y contiene " + str(count))
    else:
        logging.error("El estudio " + str(accessionNumber) + " || Modalidad: " + modality + " debe contener " + str(images) + " y sólo contiene " + str(count))


def dicom_validation(images, patient_directory, accessionNumber, modality):                                  #Mejorar la comprobación
    count = 0
    for path, dirs, files in os.walk(patient_directory):
        count = len(files)
    if images == count:
        logging.info("Los DICOM del ANC " + str(accessionNumber) + " generaron correctamente " + str(images) + "/" + str(count))
    elif count > images:
        logging.warning("El estudio " + str(accessionNumber) + " || Modalidad: " + modality + " debe contener " + str(images) + " y contiene " + str(count))
    else:
        logging.error("El estudio " + str(accessionNumber) + " || Modalidad: " + modality + " debe contener " + str(images) + " y sólo contiene " + str(count))

def main(STATE,template_directory, download_directory):
    logging.info("INIT :: Inicia la descarga ")
    for project, units in BRANCH_OFFICE.items():
        if STATE == project:
            conexionPacs = get_connection(DB_CONNECTIONS["ARC"].format(HOST=HOST))
            studyPatient = conexionPacs.cursor()
            templates = os.listdir(template_directory)
            templates.sort()
            for file in templates:
                template_file = os.path.join(template_directory,file)
                print("\nProcesando:  " + file)  
                logging.info("Procesando el archivo: " + file)
                noPatients = patients_count(template_file)
                print("Procesando: " + str(noPatients) + " pacientes")
                book = openpyxl.load_workbook(template_file, data_only=True)
                sheet = book.active
                for i in range(2, (int(noPatients))):
                    results = directory_create(sheet,i, download_directory)
                    accessionNumber = results[1]
                    patient_directory = results[0]
                    modality = results[2]
                    patient = results[3]
                    images = results[4]
                    if modality == "DXA":
                        studyPatient.execute(PACS_STUDIES_DXA.format(accessionNumber=accessionNumber))
                    else:
                        studyPatient.execute(PACS_STUDIES.format(accessionNumber=accessionNumber))
                    uid_result = str(studyPatient.fetchall())                                       #Variable que guarda el resultado de la consulta
                    UID = re.sub("[]()',[]", "", uid_result,  flags=re.IGNORECASE)
                    #----------------------------------------------DESCARGA DE ARCHIVOS DICOM MEDIANTE GETSCU-------------------------------------------------------------------
                    oldfiles = os.path.isdir(patient_directory)                  
                    if not oldfiles:
                        try:
                            subprocess.run(["python3", "-m", "pynetdicom", "getscu", "-aet", "SUIM_PACS", "-aec", AET, HOST, PORT, "-k", "StudyInstanceUID=" + UID + "", "-k", "(0008,0052)=STUDY", "-od", patient_directory + "", "-S"])
                            #python3 -m pynetdicom getscu -aet SUIM_PACS -aec SUIM 192.168.100.100 11112 -k StudyInstanceUID=2.25.337218133025685918088426135834635140998 -k "(0008,0052)=STUDY" -od . -S
                        except:
                            print("Ocurrió un Error en la descarga")
                            logging.warning("Error en descarga")
                            exit()
                    else:
                        first_validation(images, patient_directory, accessionNumber, modality)
                        continue
                    #----------------------------------------------RENOMBRADO de DICOM------------------------------------------------------------------------------------------
                    dicom_rename(UID, accessionNumber,patient_directory)
                    dicom_validation(images, patient_directory, accessionNumber, modality)
                    #----------------------------------------------CERRAR CONEXIONES--------------------------------------------------------------------------------------------
            conexionPacs.close()
    logging.info("END :: Descarga finalizada")

if __name__ == "__main__":                                                                          # Llamar a la función main si este script se ejecuta como el programa principal
    __LOG__ = "logger-" + datetime.datetime.now().strftime("%d%m-%H-%M-%S") + ".log"
    logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s : %(levelname)s : %(message)s',
                    filename = __LOG__,
                    filemode = 'w',)
    main(STATE, sys.argv[1], sys.argv[2])